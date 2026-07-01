"""
Excel 文件解析模块
"""
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple
import re
from datetime import datetime, date
import calendar
import logging
from dateutil.relativedelta import relativedelta
from app.core.ip_formatter import IPFormatter, PortFormatter

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel 解析器"""

    # 关键字段，用于查找真正的表头
    KEY_FIELDS = ["源IP", "目的IP", "目的端口"]
    FIELD_MAPPING = {
        "源端系统-环境-用途": "source_system_name",
        "源IP": "source_ip",
        "目的端系统-环境-用途": "dest_system_name",
        "目的IP": "dest_ip",
        "目的端口": "service",
        "使用时间": "usage_time",
    }

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def parse(self) -> Dict[str, Any]:
        """
        解析 Excel 文件
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)

            # 1. 优先读取名称包含"网络策略"的 sheet
            self.sheet = self._find_network_policy_sheet()
            logger.info(f"使用 sheet: {self.sheet.title}")

            # =========================================================================
            # 【阶段 1】：原始表格（刚上传表格、读取出内容时）
            # 注意：此时处于“找到表头之前”，使用 Excel 列字母（A, B, C...）作为键，从第1行起盲读全表
            # =========================================================================
            original_data = self._read_raw_upload_data()
            logger.info(f"原始表格（上传时）数据读取完成，共 {len(original_data)} 行")

            # 2. 查找真正的表头行
            header_row, headers = self._find_real_header()
            logger.info(f"表头行: {header_row}, 表头: {headers}")

            if not headers:
                raise Exception("未找到包含关键字段（源IP、目的IP、目的端口）的表头行")

            # 3. 读取数据（从真正的表头行之后开始）
            raw_data = self._read_data(headers, header_row)
            logger.info(f"按真实表头读取到 {len(raw_data)} 行数据")

            # 移除空列
            cleaned_data, valid_headers = self._remove_empty_columns(raw_data, headers)
            headers = valid_headers

            # =========================================================================
            # 【阶段 2】：第一版表格（找到列头的时候）
            # 此时移除了无关空列，且以真正的中文列头作为键，但数据尚未经过任何格式化清洗
            # =========================================================================
            v1_data = [dict(row) for row in cleaned_data]
            logger.info(f"第一版表格（找到列头）提取完成，共 {len(v1_data)} 行数据")

            # 4. 数据清洗与格式化处理
            # 格式化 IP 地址和端口
            processing_data = self._format_ip_addresses([dict(row) for row in v1_data])
            # 删除示例策略
            processing_data = self._remove_example_policies(processing_data)
            # 格式化使用时间
            processing_data = self._format_usage_time(processing_data)

            # =========================================================================
            # 【阶段 3】：第二版表格（删除示例，格式化IP、时间等的时候）
            # =========================================================================
            v2_data = [dict(row) for row in processing_data]
            logger.info(f"第二版处理完成（已删除示例并格式化IP、时间），剩余 {len(v2_data)} 行数据")

            # 5. 转换成英文字段名（用于最终保存到数据库）
            final_data = self._normalize_field_names([dict(row) for row in v2_data])
            logger.info(f"字段名标准化完成")

            if final_data:
                logger.info(f"第一行数据示例: {final_data[0]}")

            return {
                "headers": headers,
                "original_data": original_data,  # 原始表格（上传表格、找表头之前的时候，键为 A, B, C...）
                "formatted_v1_data": v1_data,              # 第一版表格（找到列头的时候，键为 中文列名）
                "formatted_v2_data": v2_data,              # 第二版表格（删除示例，格式化等之后，键为 中文列名）
                "data": final_data,              # 最终表格（英文字段名，用于入库）
                "total_rows": len(final_data),
                "header_row": header_row
            }
        except Exception as e:
            logger.error(f"Excel 解析失败: {str(e)}")
            raise Exception(f"Excel 解析失败: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()

    def _read_raw_upload_data(self) -> List[Dict[str, Any]]:
        """
        在寻找真实表头之前，直接按行读取整个 Sheet 的绝对原始数据。
        使用列字母（A, B, C...）作为字典的键，完整还原上传时的表格全貌。
        """
        data = []
        for row_idx in range(1, self.sheet.max_row + 1):
            row = list(self.sheet[row_idx])
            row_data = {}
            for col_idx, cell in enumerate(row, 1):
                col_letter = get_column_letter(col_idx)
                row_data[col_letter] = self._format_value(cell.value)

            # 只要这一行有任意一个单元格有值（不为""），就视作有效原始行
            if any(val != "" for val in row_data.values()):
                row_data["_row_number"] = row_idx
                data.append(row_data)
        return data

    def _find_network_policy_sheet(self):
        for sheet in self.workbook.worksheets:
            if "网络策略" in sheet.title:
                return sheet
        return self.workbook.active

    def _remove_empty_columns(self, data: List[Dict[str, Any]], headers: List[str]) -> Tuple[List[Dict[str, Any]], List[str]]:
        if not data or not headers:
            return data, headers

        empty_columns = []
        for col in headers:
            is_empty = True
            for row in data:
                value = row.get(col, "")
                if value not in ("", None) and str(value).strip() != "":
                    is_empty = False
                    break
            if is_empty:
                empty_columns.append(col)

        if empty_columns:
            for row in data:
                for col in empty_columns:
                    row.pop(col, None)

        new_headers = [h for h in headers if h not in empty_columns]
        return data, new_headers

    def _find_real_header(self) -> Tuple[int, List[str]]:
        for row_idx in range(1, min(self.sheet.max_row + 1, 20)):
            row = list(self.sheet[row_idx])
            row_values = [str(cell.value).strip() if cell.value else "" for cell in row]

            if self._contains_key_fields(row_values):
                logger.info(f"找到表头行: 第 {row_idx} 行")
                return row_idx, row_values

        logger.warning("未找到包含关键字段的表头行，使用第一行")
        first_row = list(self.sheet[1])
        headers = [str(cell.value).strip() if cell.value else "" for cell in first_row]
        return 1, headers

    def _contains_key_fields(self, row_values: List[str]) -> bool:
        found_fields = set()
        for value in row_values:
            value = str(value).strip()
            if len(value) < 50:
                for key_field in self.KEY_FIELDS:
                    if key_field in value:
                        found_fields.add(key_field)

        return len(found_fields) == len(self.KEY_FIELDS)

    def _read_data(self, headers: List[str], header_row: int) -> List[Dict[str, Any]]:
        data = []
        for row_idx in range(header_row + 1, self.sheet.max_row + 1):
            row = list(self.sheet[row_idx])
            row_data = {}

            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    row_data[header] = self._format_value(cell.value)

            if any(row_data.values()):
                row_data["_row_number"] = row_idx
                data.append(row_data)
        return data

    def _normalize_field_names(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        normalized_data = []
        for row in data:
            normalized_row = {}
            for key, value in row.items():
                standard_key = self.FIELD_MAPPING.get(key, key)
                normalized_row[standard_key] = value
            normalized_data.append(normalized_row)
        return normalized_data

    def _format_ip_addresses(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        for row in data:
            if row.get("源IP"):
                row["源IP"] = IPFormatter.format_ip_list(str(row["源IP"]))
            if row.get("目的IP"):
                row["目的IP"] = IPFormatter.format_ip_list(str(row["目的IP"]))
            if row.get("目的端口"):
                row["目的端口"] = PortFormatter.format_port_list(str(row["目的端口"]))
        return data

    def _remove_example_policies(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        filtered_data = []
        src_key = "源端系统-环境-用途"
        dst_key = "目的端系统-环境-用途"

        for row in data:
            source_system = str(row.get(src_key, ""))
            dest_system = str(row.get(dst_key, ""))

            if "示例" in source_system and "示例" in dest_system:
                logger.info(f"过滤示例策略: 源={source_system}, 目的={dest_system}")
                continue

            filtered_data.append(row)
        return filtered_data

    def _format_usage_time(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        time_key = "使用时间"

        for row in data:
            if time_key not in row or not row[time_key]:
                continue

            time_value = str(row[time_key]).strip()
            formatted_time = '长期'

            try:
                if '长期' in time_value:
                    formatted_time = '长期'
                elif '个月' in time_value:
                    months_match = re.search(r'(\d+)个月', time_value)
                    if months_match:
                        months = int(months_match.group(1))
                        future_date = date.today() + relativedelta(months=+months)
                        last_day = calendar.monthrange(future_date.year, future_date.month)[1]
                        formatted_time = f'{future_date.year}/{future_date.month:02d}/{last_day}'
                else:
                    date_match = re.match(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', time_value)
                    if date_match:
                        year = int(date_match.group(1))
                        month = int(date_match.group(2))
                        last_day = calendar.monthrange(year, month)[1]
                        formatted_time = f'{year}/{month:02d}/{last_day}'
                    else:
                        formatted_time = '长期'

            except Exception as e:
                logger.warning(f"时间格式化失败: {time_value}, 错误: {e}")
                formatted_time = '长期'

            row[time_key] = formatted_time

        return data

    def _format_value(self, value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, str):
            return value.strip()
        return value


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    excel_file_path = "123.xlsx"
    excel_parser = ExcelParser(excel_file_path)
    try:
        parsed_result = excel_parser.parse()
        print("\n=== 各阶段数据量核对 ===")
        print(f"原始表格 (找表头前): {len(parsed_result['original_data'])} 行 (包含表头行以及上方的标题行等)")
        print(f"第一版表格 (找到列头): {len(parsed_result['v1_data'])} 行")
        print(f"第二版表格 (清洗完毕): {len(parsed_result['v2_data'])} 行")

        print("\n=== 原始表格第一行示例 ===")
        if parsed_result['original_data']:
            print(parsed_result['original_data'][0])
    except Exception as e:
        print(f"执行出错: {e}")